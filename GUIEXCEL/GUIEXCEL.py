import tkinter as tk
from tkinter import messagebox
from tkinter import scrolledtext
import cv2
import pandas as pd
from ultralytics import YOLO
import cvzone
import threading
import os
import serial
from colorama import Fore
from geopy.distance import geodesic
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Global variables
polygon = [
    (13.8197316667, 100.5150371667),
    (13.8197428333, 100.5150236667),
    (13.8197716667, 100.5150546667),
    (13.8197738333, 100.5150331667),
]

square_x = 87
square_y = 10
square_size = 415

model = YOLO('D:/phyton/project laser/New folder/colorv8.2.pt')
center_point = []
precenter_x = 0
precenter_y = 0

ser = None
cap = None
is_detecting = False
inarea = False
data = []

def split_data(data):
    for i in data:
        print(i[0], i[1])
        save_to_file(i[0], i[1])
    print(Fore.GREEN + "Data saved to file")

def check_and_clear_file(file_path):
    if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
        open(file_path, "w").close()

def is_inside_square(x, y):
    return square_x <= x <= square_x + square_size and square_y <= y <= square_y + square_size

def save_to_file(distance_x_cm, distance_y_cm, c=""):
    with open("D:/phyton/project laser/distance_data.txt", "a") as file:
        file.write(f"{distance_x_cm:.2f} {distance_y_cm:.2f} {c},")
        log_message(f"Data saved to file: {distance_x_cm:.2f}, {distance_y_cm:.2f}")

def read_and_send_to_arduino(file_path, serial_port):
    with open(file_path, "r") as file:
        lines = file.readlines()
    log_message("Sending data to Arduino...")
    for line in lines:
        text = '#' + line.rstrip(',') + ';' + '\n'
        print(text)
        serial_port.write(text.encode())
    log_message("Data sent to Arduino")

def start_detection_thread():
    detection_thread = threading.Thread(target=start_detection)
    detection_thread.start()

def stop_detection():
    global is_detecting
    is_detecting = False
    log_message("Detection stopped")

def start_detection_harr_thread():
    detection_thread = threading.Thread(target=start_detection_harrcascade)
    detection_thread.start()

def save_data():
    if data:
        for d in data:
            save_to_file(d[0], d[1])
        print(Fore.GREEN + "Data saved to file")
        messagebox.showinfo("Information", "Data saved to file")
    else:
        messagebox.showwarning("Warning", "No data to save")

def clear_data():
    check_and_clear_file("D:/phyton/project laser/distance_data.txt")
    log_message("Data cleared from file")
    messagebox.showinfo("Information", "Data cleared from file")

def send_data():
    read_and_send_to_arduino("D:/phyton/project laser/distance_data.txt", ser)
    check_and_clear_file("D:/phyton/project laser/distance_data.txt")
    log_message("Data sent to Arduino")
    messagebox.showinfo("Information", "Data sent to Arduino")

def save_to_excel():
    global data
    try:
        # Create a new Excel file if it doesn't exist, otherwise load the existing one
        excel_file = "D:/phyton/project laser/Book2.xlsx"
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Detection Data"
            ws.append(["Time", "Latitude", "Longitude", "Distance X (cm)", "Distance Y (cm)","Class"])
        else:
            wb = load_workbook(excel_file)
            ws = wb.active
        
        # Append the latest data
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        latitude, longitude = get_current_lat_lon()  # Implement this function to get current lat/lon
        if latitude is None or longitude is None:
            messagebox.showwarning("Warning", "Unable to get current GPS data. Make sure the GPS device is connected.")
            return
        print("data :",data)
        for d in data:
           ws.append([now, latitude, longitude, d[0], d[1],d[2]])
        # Save the workbook
        wb.save(excel_file)
        log_message("Data saved to Excel")
        data = []
        messagebox.showinfo("Information", "Data saved to Excel")
    except Exception as e:
        log_message(f"Error saving to Excel: {e}")
        messagebox.showerror("Error", f"Error saving to Excel: {e}")

def start_detection():
    global cap, ser, is_detecting, precenter_x, precenter_y, center_point, data, inarea
    is_detecting = True
    log_message("Starting detection...")
    ser = serial.Serial('COM13', 9600) #arduino
    cap = cv2.VideoCapture(0)

    check_and_clear_file("D:/phyton/project laser/distance_data.txt")
    log_message("Cleared previous data from file")

    with open("D:/phyton/project laser/cocoforv8.txt", "r") as my_file:
        class_list = my_file.read().split("\n")

    count = 0
    data = []
    while is_detecting:
        ret, frame = cap.read()
        if not ret:
            break

        count += 1
        if count % 5 != 0:
            continue
        frame = cv2.resize(frame, (600, 400))
        results = model.predict(frame)
        if results:
            a = results[0].boxes.data
            px = pd.DataFrame(a).astype("float")
            #data = []
            for index, row in px.iterrows():
                x1, y1, x2, y2, _, d = map(int, row[:6])
                
                if d < len(class_list):
                    c = class_list[d]
                    log_message(f"class ID is {c}")
                else:
                    log_message(f"Invalid class ID {d}")
                    continue

                center_x = int((x1 + x2) / 2)
                center_y = int((y1 + y2) / 2)

                if center_x != precenter_x or center_y != precenter_y:
                    precenter_x = center_x
                    precenter_y = center_y
                
                    log_message(f"inarea is {inarea}")
                    if is_inside_square(center_x, center_y) and inarea == 0: # ถ้าไม่อยู่ในพื้นที่และต้องการ detect 0 
                        cv2.rectangle(frame, (x1, y1), (x2, y2), (100, 0, 100), 1)
                        cv2.circle(frame, (center_x, center_y), 3, (0, 0, 255), -1)
                        cvzone.putTextRect(frame, f'{c}', (x1, y1), 0.6, 1)

                        distance_x_cm = (-4E-06 * center_x ** 2 + 0.0408 * center_x - 3.1945) + 0.4 + 0.6
                        distance_y_cm = (2E-06 * center_y ** 2 - 0.0452 * center_y + 17.705) + 1.5 - 1 + 0.2
                        
                        if c != "kana":
                            data.append([round(distance_x_cm, 2), round(distance_y_cm, 2),c])
                            if len(data) > len(px):
                                data.pop()
            
        cv2.imshow("Weed detect", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    ser.close()
    cap.release()
    cv2.destroyAllWindows()
    log_message("Detection stopped")

def start_detection_harrcascade():
    cap = cv2.VideoCapture(0)
    detectobject = cv2.CascadeClassifier('D:/phyton/project laser/New folder/kana new.xml')

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = detectobject.detectMultiScale(gray, 1.3, 5)
        count = 0

        for (x, y, w, h) in faces:
            cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 100, 59), 2)

            center_x = x + w // 2
            center_y = y + h // 2
            cv2.circle(frame, (center_x, center_y), 5, (0, 0, 255), -1)

            distance_x_cm = (-4E-06 * center_x ** 2 + 0.0408 * center_x - 3.1945) + 0.4 + 0.6
            distance_y_cm = (2E-06 * center_y ** 2 - 0.0452 * center_y + 17.705) + 1.5 - 1 + 0.2

            center_point.append((distance_x_cm, distance_y_cm))

            if len(center_point) >= 2:
                if center_point[-2] != center_point[-1]:
                    log_message(f"Distance X: {distance_x_cm:.2f} cm, Distance Y: {distance_y_cm:.2f} cm")
                    save_to_file(distance_x_cm, distance_y_cm)
                    count += 1

        cv2.imshow('frame', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

def stop_code():
    root.quit()

def log_message(message):
    text_box.insert(tk.END, message + '\n')
    text_box.see(tk.END)

def clear_log():
    text_box.delete('1.0', tk.END)

def get_current_lat_lon():
    try:
        ser_gprmc = serial.Serial('COM5', 9600)
        while True:
            line = ser_gprmc.readline().decode('ascii', errors='replace')
            if line.startswith('$GPRMC'):
                time, lat, lon = parse_gprmc(line)
                if time and lat and lon:
                    return lat, lon
    except serial.SerialException as e:
        log_message(f"Serial error: {e}")
    except Exception as e:
        log_message(f"Error: {e}")
    return None, None

def parse_gprmc(gprmc_data):
    try:
        parts = gprmc_data.split(',')
        if parts[0] == '$GPRMC' and parts[2] == 'A':  # Only process valid data
            time_str = parts[1]
            lat_str = parts[3]
            lat_dir = parts[4]
            lon_str = parts[5]
            lon_dir = parts[6]

            time = f"{time_str[0:2]}:{time_str[2:4]}:{time_str[4:6]}"

            lat = float(lat_str[:2]) + float(lat_str[2:]) / 60
            if lat_dir == 'S':
                lat = -lat

            lon = float(lon_str[:3]) + float(lon_str[3:]) / 60
            if lon_dir == 'W':
                lon = -lon

            return time, lat, lon
    except Exception as e:
        log_message(f"Error parsing GPRMC data: {e}")
    return None, None, None

def read_gprmc():
    global inarea
    try:
        ser_gprmc = serial.Serial('COM5', 9600)
        while True:
            line = ser_gprmc.readline().decode('ascii', errors='replace')
            if line.startswith('$GPRMC'):
                time, lat, lon = parse_gprmc(line)
                if time and lat and lon:
                    log_message(f"Time: {time}, Latitude: {lat:.6f}, Longitude: {lon:.6f}")
                    inarea = False
                    inarea = point_inside_polygon(lat,lon,polygon) # ADD IN AREA
                    log_message(f"in area is {inarea}")
                    messagebox.showinfo("GPRMC Data", f"Time: {time}\nLatitude: {lat:.6f}\nLongitude: {lon:.6f}")
                    break
    except serial.SerialException as e:
        log_message(f"Serial error: {e}")
    except Exception as e:
        log_message(f"Error: {e}")

def start_read_gprmc_thread():
    gprmc_thread = threading.Thread(target=read_gprmc)
    gprmc_thread.start()

def point_inside_polygon(x=0, y=0, poly=None):
    n = len(poly)
    inside = False
    p1x, p1y = poly[0]
    for i in range(n + 1):
        p2x, p2y = poly[i % n]
        if y > min(p1y, p2y):
            if y <= max(p1y, p2y):
                if x <= max(p1x, p2x):
                    if p1y != p2y:
                        xinters = (y - p1y) * (p2x - p1x) / (p2y - p1y) + p1x
                    if p1x == p2x or x <= xinters:
                        inside = not inside
        p1x, p1y = p2x, p2y
    return inside

# GUI setup
root = tk.Tk()
root.title("Detection GUI")
root.geometry("380x460")

start_button = tk.Button(root, text="YOLO DETECT", command=start_detection_thread, bg="#4CAF50", fg="black", activebackground="#45a049", activeforeground="white")
start_button.place(x=50, y=50, width=120, height=30)

auto_detect_button = tk.Button(root, text="HARRCASCADE", command=start_detection_harr_thread, bg="#311b92", fg="white", activebackground="#7e57c2", activeforeground="white")
auto_detect_button.place(x=200, y=50, width=120, height=30)

stop_button = tk.Button(root, text="หยุดตรวจจับ", command=stop_detection, bg="#bf360c", fg="black", activebackground="#7e57c2", activeforeground="white")
stop_button.place(x=50, y=100, width=120, height=30)

save_button = tk.Button(root, text="บันทึกข้อมูล", command=save_data, bg="#546e7a", fg="white", activebackground="#263238", activeforeground="red")
save_button.place(x=200, y=100, width=120, height=30)

clear_button = tk.Button(root, text="ล้างข้อมูล", command=clear_data, bg="#ffeb3b", fg="black", activebackground="#f57f17", activeforeground="red")
clear_button.place(x=50, y=150, width=120, height=30)

send_button = tk.Button(root, text="ส่งออกข้อมูล", command=send_data, bg="#9e9d24", fg="black", activebackground="#e6ee9c", activeforeground="red")
send_button.place(x=200, y=150, width=120, height=30)

gprmc_button = tk.Button(root, text="อ่านค่า GPRMC", command=start_read_gprmc_thread, bg="#607d8b", fg="white", activebackground="#455a64", activeforeground="red")
gprmc_button.place(x=125, y=200, width=120, height=30)

excel_button = tk.Button(root, text="บันทึกลง Excel", command=save_to_excel, bg="#03A9F4", fg="white", activebackground="#0288D1", activeforeground="white")
excel_button.place(x=125, y=250, width=120, height=30)

text_box = scrolledtext.ScrolledText(root, wrap=tk.WORD, width=35, height=6)
text_box.grid(row=10, column=0, columnspan=2, padx=20, pady=10)
text_box.place(x=50, y=300)

clear_log_button = tk.Button(root, text="Clear Log", command=clear_log)
clear_log_button.place(x=200, y=410, width=120, height=30)

exit_button = tk.Button(root, text="Exit", command=stop_code or stop_detection)
exit_button.place(x=50, y=410, width=120, height=30)
root.mainloop()
